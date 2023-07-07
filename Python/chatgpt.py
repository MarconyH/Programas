import sys
import os
import openpyxl
from PIL import Image

# Obtendo o caminho do diretório onde o script está localizado
script_path = os.path.dirname(os.path.realpath(sys.executable))

# Concatenando com o nome da pasta de imagens
folder_path = os.path.join(script_path, 'imagens')
print(folder_path)
if not os.path.exists(folder_path):
    print("Pasta de imagens não encontrada.")
    sys.exit()

# Concatenando com o nome do arquivo da planilha
workbook_path = os.path.join(script_path, 'imagens.xlsx')

# Verificando se a planilha já existe
if os.path.exists(workbook_path):
    # Abrindo a planilha existente
    workbook = openpyxl.load_workbook(workbook_path)
    sheet = workbook['Sheet']

    # Contador de linhas
    row_count = sheet.max_row + 1

    # Lista de nomes de imagens já adicionadas
    added_images = [cell.value for cell in sheet['A']]
else:
    # Criando uma nova planilha do Excel
    workbook = openpyxl.Workbook()
    sheet = workbook.active

    # Adicionando títulos às colunas
    sheet.cell(row=1, column=1).value = 'Nome'
    sheet.cell(row=1, column=2).value = ''

    # Contador de linhas
    row_count = 2

    # Lista de nomes de imagens já adicionadas
    added_images = []
    
    # Percorrendo as imagens na pasta
    for image_name in os.listdir(folder_path):
        image_path = os.path.join(folder_path, image_name)
        try:
            with Image.open(image_path) as img:
                # Removendo extensão do nome do arquivo
                image_name = os.path.splitext(image_name)[0]

                # Extraindo tamanho do nome do arquivo
                size = image_name.split(" ")[-2]+" "+image_name.split(" ")[-1]

                # Verificando se a imagem já foi adicionada
                if image_name not in added_images:
                    # Adicionando informações à planilha
                    sheet.cell(row=row_count, column=1).value = image_name
                    sheet.cell(row=row_count, column=2).value = size
                    # Adicionando a imagem na planilha
                    sheet.column_dimensions[openpyxl.utils.get_column_letter(3)].width = 30
                    sheet.row_dimensions[row_count].height = 100
                    img = openpyxl.drawing.image.Image(image_path)
                    img.width = 100
                    img.height = 100
                    img.anchor = openpyxl.utils.get_column_letter(3) + str(row_count)
                    sheet.add_image(img)
                    added_images.append(image_name)
                    # Incrementando o contador de linhas
                    row_count += 1
        except:
            print(f"Error reading image {image_path}")
    # Salvando a planilha
    workbook.save(workbook_path)

    # Mensagem de confirmação
    input("Processo concluído com sucesso! Pressione enter para sair.")

